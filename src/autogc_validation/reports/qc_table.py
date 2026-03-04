# -*- coding: utf-8 -*-
"""
QC Review table builder and Excel writer for the MDVR spreadsheet.

Converts the wide boolean failure DataFrames produced by compounds_above_mdl
into a human-readable table and writes it to the 'QC Review' sheet of the
MDVR Excel template.
"""

import pandas as pd
from openpyxl import load_workbook

from autogc_validation.database.enums import (
    ColumnType,
    PLOT_CODES,
    BP_CODES,
    COLUMN_CALIBRANTS,
    aqs_to_name,
)

_BLANK_ACTIONS = (
    "Compounds above their respective MDLs qualified with flag LB forward and "
    "backward to the nearest passing blank. Compounds above 0.5 ppbC nulled with "
    "flag AS forward and backward to the nearest passing blank."
)

_RTS_ACTIONS = "Failures noted, no data qualification performed."

_PLOT_CALIBRANT = COLUMN_CALIBRANTS[ColumnType.PLOT]
_BP_CALIBRANT   = COLUMN_CALIBRANTS[ColumnType.BP]


def _build_notes(mdl_row: pd.Series, thresh_row: pd.Series, codes: list[int]) -> str:
    """Build the notes cell text for a set of compound codes.

    Args:
        mdl_row: Boolean Series for one sample row (1 = above MDL).
        thresh_row: Boolean Series for one sample row (1 = above threshold).
        codes: AQS codes to check (e.g. all PLOT_CODES present in the DataFrame).

    Returns:
        Multi-line string, e.g.:
            "Compounds above MDL: Benzene, Toluene\\nCompounds above 0.5 ppbC: Ethylbenzene"
        Empty string if no codes are flagged.
    """
    mdl_names = [aqs_to_name(c) for c in codes if mdl_row.get(c, 0) == 1]
    thresh_names = [aqs_to_name(c) for c in codes if thresh_row.get(c, 0) == 1]

    lines = []
    if mdl_names:
        lines.append(f"Compounds above MDL: {', '.join(mdl_names)}")
    if thresh_names:
        lines.append(f"Compounds above 0.5 ppbC: {', '.join(thresh_names)}")
    return "\n".join(lines)


def build_blank_qc_table(
    mdl_failures: pd.DataFrame,
    threshold_failures: pd.DataFrame,
) -> pd.DataFrame:
    """Build a blank QC summary table for the MDVR QC Review sheet.

    Converts the two wide boolean failure DataFrames from compounds_above_mdl
    into one row per blank sample with human-readable notes columns.

    Args:
        mdl_failures: Wide boolean DataFrame — 1 where compound exceeded its MDL.
            Columns: filename + integer AQS codes. Index name: date_time.
        threshold_failures: Wide boolean DataFrame — 1 where compound exceeded
            0.5 ppbC threshold. Same shape and index as mdl_failures.

    Returns:
        DataFrame with columns:
            date       MM/DD/YYYY formatted string
            time       HH:00 formatted string
            filename   original sample filename
            plot_notes "Compounds above MDL: ...\\nCompounds above 0.5 ppbC: ..."
                       (empty string if no PLOT failures)
            bp_notes   same format for BP-column compounds
            actions    blank actions message if any failure; "None taken." otherwise
    """
    compound_cols = [c for c in mdl_failures.columns if isinstance(c, int)]
    plot_cols = [c for c in compound_cols if c in PLOT_CODES]
    bp_cols = [c for c in compound_cols if c in BP_CODES]

    rows = []
    for i, (timestamp, mdl_row) in enumerate(mdl_failures.iterrows()):
        thresh_row = threshold_failures.iloc[i]

        plot_notes = _build_notes(mdl_row, thresh_row, plot_cols)
        bp_notes = _build_notes(mdl_row, thresh_row, bp_cols)

        any_failure = bool(
            (mdl_row[compound_cols] == 1).any()
            or (thresh_row[compound_cols] == 1).any()
        )
        actions = _BLANK_ACTIONS if any_failure else "None taken."

        rows.append({
            "date": timestamp.strftime("%m/%d/%Y"),
            "time": timestamp.strftime("%H:00"),
            "filename": mdl_row["filename"],
            "plot_notes": plot_notes,
            "bp_notes": bp_notes,
            "actions": actions,
        })

    return pd.DataFrame(rows)


_PRECISION_ACTIONS = (
    "Qualified failing compound(s) with flag QX forward and backward "
    "to nearest passing CVS precision."
)

_VALID_QC_TYPES = {"CVS", "LCS", "RTS"}


def _build_precision_notes(fail_row: pd.Series, codes: list[int]) -> str:
    """Build notes for a CVS precision failure row.

    Args:
        fail_row: Integer Series for one sample row (1 = RPD exceeded, 0 = pass).
        codes: AQS codes to check (e.g. all PLOT_CODES present in the DataFrame).

    Returns:
        "Failing compounds: name1, name2" or empty string if none flagged.
    """
    names = [aqs_to_name(c) for c in codes if fail_row.get(c, 0) == 1]
    if not names:
        return ""
    return f"Failing compounds: {', '.join(names)}"


def build_precision_qc_table(
    precision_failures: pd.DataFrame,
) -> pd.DataFrame:
    """Build a CVS precision QC summary table for the MDVR QC Review sheet.

    One row per back-to-back CVS pair, identified by the first run's timestamp.

    Args:
        precision_failures: DataFrame from check_cvs_precision.
            Index: first-run timestamp. Columns: 'filename' + integer AQS codes.
            Values: 1 if RPD exceeded threshold, 0 otherwise.

    Returns:
        DataFrame with columns:
            date       MM/DD/YYYY formatted string
            time       HH:00 formatted string
            filename   filename of the first run in the pair
            plot_notes "Failing compounds: name1, name2" or empty string
            bp_notes   same for BP-column compounds
            actions    precision actions message or "None taken."
    """
    compound_cols = [c for c in precision_failures.columns if isinstance(c, int)]
    plot_cols = [c for c in compound_cols if c in PLOT_CODES]
    bp_cols = [c for c in compound_cols if c in BP_CODES]

    rows = []
    for timestamp, fail_row in precision_failures.iterrows():
        plot_notes = _build_precision_notes(fail_row, plot_cols)
        bp_notes = _build_precision_notes(fail_row, bp_cols)
        any_failure = any(fail_row.get(c, 0) == 1 for c in compound_cols)
        actions = _PRECISION_ACTIONS if any_failure else "None taken."

        rows.append({
            "date": timestamp.strftime("%m/%d/%Y"),
            "time": timestamp.strftime("%H:00"),
            "filename": fail_row["filename"],
            "plot_notes": plot_notes,
            "bp_notes": bp_notes,
            "actions": actions,
        })

    return pd.DataFrame(rows)


def _build_recovery_actions(
    fail_row: pd.Series,
    compound_cols: list[int],
    qc_type: str,
) -> str:
    """Build the actions cell text for one recovery QC row.

    When a calibrant species fails, a whole-column sentence is emitted
    (QX, LL or QX, LK). Individual non-calibrant failures emit a plain QX
    sentence. RTS always returns a fixed note-only message.
    """
    if qc_type == "RTS":
        return _RTS_ACTIONS

    parts = []
    plot_cal_val = fail_row.get(_PLOT_CALIBRANT, 0)
    bp_cal_val   = fail_row.get(_BP_CALIBRANT,   0)

    # Whole-column calibrant sentences.
    for cal_val, col_label, flag_code in [
        (plot_cal_val, "PLOT", "LL" if plot_cal_val == -1 else "LK"),
        (bp_cal_val,   "BP",   "LL" if bp_cal_val   == -1 else "LK"),
    ]:
        if cal_val != 0:
            parts.append(
                f"All {col_label} compounds qualified with flags QX, {flag_code} "
                f"forward and backward to the nearest passing {qc_type}."
            )

    # Individual failures on columns whose calibrant passed.
    has_individual = any(
        fail_row.get(c, 0) != 0
        and not (c in PLOT_CODES and plot_cal_val != 0)
        and not (c in BP_CODES  and bp_cal_val   != 0)
        for c in compound_cols
    )
    if has_individual:
        parts.append(
            f"Failing compound(s) qualified with flag QX forward and backward "
            f"to the nearest passing {qc_type}."
        )

    return " ".join(parts) if parts else "None taken."


def _build_recovery_notes(fail_row: pd.Series, codes: list[int]) -> str:
    """Build the notes cell text for a recovery failure row.

    Args:
        fail_row: Signed integer Series for one sample row (+1 = high failure,
            -1 = low failure, 0 = pass).
        codes: AQS codes to check (e.g. all PLOT_CODES present in the DataFrame).

    Returns:
        "Failing compounds: name1 (H), name2 (L)" or empty string if none flagged.
    """
    names = []
    for c in codes:
        val = fail_row.get(c, 0)
        if val == 1:
            names.append(f"{aqs_to_name(c)} (H)")
        elif val == -1:
            names.append(f"{aqs_to_name(c)} (L)")
    if not names:
        return ""
    return f"Failing compounds: {', '.join(names)}"


def build_recovery_qc_table(
    recovery_failures: pd.DataFrame,
    qc_type: str,
) -> pd.DataFrame:
    """Build a recovery QC summary table for the MDVR QC Review sheet.

    Converts the wide boolean failure DataFrame from check_qc_recovery into
    one row per sample with human-readable notes and a type-appropriate
    actions message.

    Args:
        recovery_failures: Wide integer DataFrame — +1 (high) or -1 (low)
            where compound recovery was outside bounds, 0 for passing samples.
            Columns: filename + integer AQS codes. Index name: date_time.
            As returned by check_qc_recovery.
        qc_type: One of 'CVS', 'LCS', or 'RTS'. Determines the actions message.

    Returns:
        DataFrame with columns:
            date       MM/DD/YYYY formatted string
            time       HH:00 formatted string
            filename   original sample filename
            plot_notes "Failing compounds: name1, name2" for PLOT-column failures
                       (empty string if none)
            bp_notes   same format for BP-column failures
            actions    calibrant-aware qualification sentence(s), or "None taken."

    Raises:
        ValueError: If qc_type is not 'CVS', 'LCS', or 'RTS'.
    """
    if qc_type not in _VALID_QC_TYPES:
        raise ValueError(
            f"qc_type must be one of {sorted(_VALID_QC_TYPES)}, got {qc_type!r}"
        )

    compound_cols = [c for c in recovery_failures.columns if isinstance(c, int)]
    plot_cols = [c for c in compound_cols if c in PLOT_CODES]
    bp_cols = [c for c in compound_cols if c in BP_CODES]

    rows = []
    for timestamp, fail_row in recovery_failures.iterrows():
        plot_notes = _build_recovery_notes(fail_row, plot_cols)
        bp_notes = _build_recovery_notes(fail_row, bp_cols)
        actions = _build_recovery_actions(fail_row, compound_cols, qc_type)

        rows.append({
            "date": timestamp.strftime("%m/%d/%Y"),
            "time": timestamp.strftime("%H:00"),
            "filename": fail_row["filename"],
            "plot_notes": plot_notes,
            "bp_notes": bp_notes,
            "actions": actions,
        })

    return pd.DataFrame(rows)


def write_qc_table_to_excel(
    table_df: pd.DataFrame,
    template_path: str,
    output_path: str,
    qc_type: str,
    start_row: int = 7,
) -> None:
    """Write a QC summary table to the 'QC Review' sheet of the MDVR template.

    Writes the qc_type label to column A at start_row, then writes
    date / time / filename / plot_notes / bp_notes / actions to columns B-G
    for each row in table_df.

    Column layout (matches MDVR template):
        A  QC Type
        B  Date (MM/DD/YYYY)
        C  Time (HH:00)
        D  Filename
        E  Plot Column Notes
        F  BP Notes
        G  Actions

    Args:
        table_df: DataFrame from build_blank_qc_table (or equivalent).
            Must have columns: date, time, filename, plot_notes, bp_notes, actions.
        template_path: Path to the MDVR Excel template (.xlsx).
        output_path: Path where the filled workbook will be saved.
            May be the same as template_path to edit in place.
        qc_type: Label written to column A at start_row
            (e.g. 'Field Blank', 'CVS', 'LCS', 'RTS').
        start_row: First data row (1-indexed). Default 7.
    """
    wb = load_workbook(template_path)
    ws = wb["QC Review"]

    ws.cell(row=start_row, column=1).value = qc_type

    for i, row in enumerate(table_df.itertuples(index=False)):
        r = start_row + i
        ws.cell(row=r, column=2).value = row.date
        ws.cell(row=r, column=3).value = row.time
        ws.cell(row=r, column=4).value = row.filename
        ws.cell(row=r, column=5).value = row.plot_notes
        ws.cell(row=r, column=6).value = row.bp_notes
        ws.cell(row=r, column=7).value = row.actions

    wb.save(output_path)
