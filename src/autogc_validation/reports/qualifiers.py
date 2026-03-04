# -*- coding: utf-8 -*-
"""
MDVR qualifier interval computation and Excel export.

Computes pass/fail intervals from blank and QC failure DataFrames,
combines compounds that share the same interval, and writes the
result to the QUALIFIERS_NULL sheet of the MDVR Excel template.
"""

import logging
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import load_workbook

from autogc_validation.database.enums import (
    ColumnType,
    PLOT_CODES,
    BP_CODES,
    COLUMN_CALIBRANTS,
    aqs_to_name,
)

logger = logging.getLogger(__name__)

_QC_TYPE_NAMES = {
    "c": "CVS",
    "e": "LCS",
    "q": "RTS",
}

_PLOT_CALIBRANT = COLUMN_CALIBRANTS[ColumnType.PLOT]
_BP_CALIBRANT   = COLUMN_CALIBRANTS[ColumnType.BP]

_MDVR_COLUMNS = [
    "Parameter(s)",
    "COMPOUND(S) or WHOLE HOUR(S) - REASON",
    "CODE",
    "startdate",
    "starthour",
    "-",
    "enddate",
    "endhour",
    "Justification",
]


def compute_failure_intervals(
    all_data: pd.DataFrame,
    failure_series: pd.Series,
    prior_sample: pd.Timestamp | None = None,
    next_sample: pd.Timestamp | None = None,
) -> np.ndarray:
    """Compute merged failure intervals from a binary pass/fail series.

    For each contiguous block of failures the interval spans from the nearest
    preceding passing observation to the nearest following passing observation.
    Overlapping intervals are merged. The ±1-hour inward shift that excludes
    the bounding passing samples is applied later in _shift_and_combine.

    When a failure block has no prior passing observation within the series,
    ``prior_sample`` is used as the left bound (e.g. the last blank from the
    previous month). When there is no following passing observation,
    ``next_sample`` is used as the right bound. If either sentinel is None the
    dataset's own min/max timestamp is used as a fallback.

    Args:
        all_data: Full Dataset.data DataFrame. Used for fallback min/max
            bounds only — not filtered or modified.
        failure_series: Series with datetime index and 0/1 values (1 = fail).
        prior_sample: Timestamp of the nearest QC/blank sample from the
            preceding month. Used as the left bound when the first sample of
            the month fails. Optional.
        next_sample: Timestamp of the nearest QC/blank sample from the
            following month. Used as the right bound when the last sample of
            the month fails. Optional.

    Returns:
        Nx2 numpy array of [left_pass_time, right_pass_time] for each
        merged failure interval.
    """
    failure_series = failure_series.copy().sort_index()
    idx = failure_series.index

    lower_bound = prior_sample if prior_sample is not None else all_data.index.min()
    upper_bound = next_sample if next_sample is not None else all_data.index.max()

    passes = failure_series == 0
    fails = failure_series != 0

    if not fails.any():
        return np.empty((0, 2), dtype="datetime64[ns]")

    pass_idx = pd.Series(idx.where(passes, pd.NaT), index=idx)

    left_bounds = pass_idx.ffill().fillna(lower_bound)
    right_bounds = pass_idx.bfill().fillna(upper_bound)

    lb = left_bounds[fails].to_numpy()
    rb = right_bounds[fails].to_numpy()

    intervals = np.column_stack([lb, rb])
    intervals = intervals[np.argsort(intervals[:, 0])]

    lb = intervals[:, 0]
    rb = intervals[:, 1]

    # Merge overlapping intervals
    new_interval_mask = np.r_[True, lb[1:] > np.maximum.accumulate(rb[:-1])]
    interval_ids = np.cumsum(new_interval_mask) - 1

    merged = np.zeros((interval_ids.max() + 1, 2), dtype=intervals.dtype)
    merged[:, 0] = np.minimum.reduceat(lb, np.flatnonzero(new_interval_mask))
    merged[:, 1] = np.maximum.reduceat(rb, np.flatnonzero(new_interval_mask))

    return merged


def _shift_and_combine(df: pd.DataFrame) -> pd.DataFrame:
    """Group qualifier rows by interval and code, combine compounds, shift times.

    Compounds sharing the same failure interval and qualifier code are combined
    into a single row. The interval is then shifted inward by one hour on each
    side to exclude the bounding passing samples from the flagged range.
    """
    df_combined = (
        df.groupby(["startdate", "starthour", "enddate", "endhour", "CODE"])
        .agg({
            "Parameter(s)": lambda x: ", ".join(x),
            "COMPOUND(S) or WHOLE HOUR(S) - REASON": "first",
            "-": "first",
            "Justification": "first",
        })
        .reset_index()
    )

    df_combined["start_dt"] = pd.to_datetime(
        df_combined["startdate"] + " " + df_combined["starthour"],
        format="%m/%d/%Y %H:%M",
    )
    df_combined["end_dt"] = pd.to_datetime(
        df_combined["enddate"] + " " + df_combined["endhour"],
        format="%m/%d/%Y %H:%M",
    )

    df_combined["start_dt"] += pd.Timedelta(hours=1)
    df_combined["end_dt"] -= pd.Timedelta(hours=1)

    df_combined["startdate"] = df_combined["start_dt"].dt.strftime("%m/%d/%Y")
    df_combined["starthour"] = df_combined["start_dt"].dt.strftime("%H:00")
    df_combined["enddate"] = df_combined["end_dt"].dt.strftime("%m/%d/%Y")
    df_combined["endhour"] = df_combined["end_dt"].dt.strftime("%H:00")

    return df_combined.drop(columns=["start_dt", "end_dt"])[_MDVR_COLUMNS]


def _make_row(code: int, reason: str, flag: str, start, end) -> dict:
    start = pd.Timestamp(start)
    end = pd.Timestamp(end)
    return {
        "Parameter(s)": aqs_to_name(code),
        "COMPOUND(S) or WHOLE HOUR(S) - REASON": reason,
        "CODE": flag,
        "startdate": start.strftime("%m/%d/%Y"),
        "starthour": start.strftime("%H:00"),
        "-": "-",
        "enddate": end.strftime("%m/%d/%Y"),
        "endhour": end.strftime("%H:00"),
        "Justification": reason,
    }


def _make_col_row(label: str, reason: str, flag: str, start, end) -> dict:
    """Like _make_row but accepts a free-text label instead of an AQS code."""
    start = pd.Timestamp(start)
    end = pd.Timestamp(end)
    return {
        "Parameter(s)": label,
        "COMPOUND(S) or WHOLE HOUR(S) - REASON": reason,
        "CODE": flag,
        "startdate": start.strftime("%m/%d/%Y"),
        "starthour": start.strftime("%H:00"),
        "-": "-",
        "enddate": end.strftime("%m/%d/%Y"),
        "endhour": end.strftime("%H:00"),
        "Justification": reason,
    }


def build_blank_qualifier_lines(
    all_data: pd.DataFrame,
    mdl_failures: pd.DataFrame,
    threshold_failures: pd.DataFrame,
    prior_blank: pd.Timestamp | None = None,
    next_blank: pd.Timestamp | None = None,
) -> pd.DataFrame:
    """Build MDVR qualifier lines for blank failures.

    MDL failures (flag LB) and threshold failures (flag AS) are handled
    separately so they receive the correct code. Compounds sharing the same
    interval and code are combined into one row.

    Args:
        all_data: Full Dataset.data DataFrame (for fallback interval bounds).
        mdl_failures: Wide boolean DataFrame — 1 where compound exceeded its
            MDL. Columns: filename + integer AQS codes. From compounds_above_mdl.
        threshold_failures: Wide boolean DataFrame — 1 where compound exceeded
            0.5 ppbC. Same shape as mdl_failures.
        prior_blank: Timestamp of the last blank sample from the preceding
            month. Used as the left bound when the first blank of the month
            fails. Optional.
        next_blank: Timestamp of the first blank sample from the following
            month. Used as the right bound when the last blank of the month
            fails. Optional.

    Returns:
        DataFrame with MDVR qualifier columns ready for Excel export.
    """
    compound_cols = [c for c in mdl_failures.columns if isinstance(c, int)]
    rows = []

    for code in compound_cols:
        merged = compute_failure_intervals(
            all_data, mdl_failures[code], prior_blank, next_blank
        )
        for start, end in merged:
            rows.append(_make_row(
                code, "Blank(s) above respective MDL(s)", "LB", start, end
            ))

    thresh_cols = [c for c in threshold_failures.columns if isinstance(c, int)]
    for code in thresh_cols:
        merged = compute_failure_intervals(
            all_data, threshold_failures[code], prior_blank, next_blank
        )
        for start, end in merged:
            rows.append(_make_row(
                code, "Blank(s) above 0.5 ppbC threshold", "AS", start, end
            ))

    if not rows:
        return pd.DataFrame(columns=_MDVR_COLUMNS)

    return _shift_and_combine(pd.DataFrame(rows))


def build_qc_qualifier_lines(
    all_data: pd.DataFrame,
    qc_failures: pd.DataFrame,
    qc_type: str,
    prior_qc: pd.Timestamp | None = None,
    next_qc: pd.Timestamp | None = None,
) -> pd.DataFrame:
    """Build MDVR qualifier lines for QC recovery failures.

    RTS failures are noted but not data-qualified — returns an empty DataFrame
    for qc_type 'q'. CVS ('c') and LCS ('e') failures are qualified as follows:

    - If the PLOT-column calibrant (Propane) failed on a run, all PLOT-column
      compounds receive flag **LL** (calibrant low) or **LK** (calibrant high)
      for the interval spanning that run.
    - If the BP-column calibrant (Toluene) failed on a run, all BP-column
      compounds receive LL or LK similarly.
    - When a calibrant fails, a single qualifier row is written with
      ``Parameter(s)`` = "All PLOT compounds" or "All BP compounds" and
      ``CODE`` = "QX, LL" (low) or "QX, LK" (high).
    - Compounds whose column calibrant passed but which individually failed
      receive **QX** (one row per compound, grouped by shared interval).

    Compounds sharing the same failure interval and flag code are combined into
    one qualifier row by ``_shift_and_combine``.

    Args:
        all_data: Full Dataset.data DataFrame.
        qc_failures: Wide integer DataFrame — +1 (high) or -1 (low) where
            compound recovery was outside bounds, 0 for passing samples.
            Columns: filename + integer AQS codes. From check_qc_recovery.
        qc_type: Sample type code — 'c' (CVS), 'e' (LCS), or 'q' (RTS).
        prior_qc: Timestamp of the last QC sample from the preceding month.
            Used as the left bound when the first QC sample of the month fails.
            Optional.
        next_qc: Timestamp of the first QC sample from the following month.
            Used as the right bound when the last QC sample of the month fails.
            Optional.

    Returns:
        DataFrame with MDVR qualifier columns, or empty DataFrame for RTS.
    """
    if qc_type == "q":
        logger.info("RTS failures noted; no data qualification performed")
        return pd.DataFrame(columns=_MDVR_COLUMNS)

    qc_name = _QC_TYPE_NAMES[qc_type]
    compound_cols = [c for c in qc_failures.columns if isinstance(c, int)]
    plot_cols = [c for c in compound_cols if c in PLOT_CODES]
    bp_cols   = [c for c in compound_cols if c in BP_CODES]

    # Calibrant failure series for each column (zero series if calibrant absent).
    plot_cal = (
        qc_failures[_PLOT_CALIBRANT]
        if _PLOT_CALIBRANT in compound_cols
        else pd.Series(0, index=qc_failures.index)
    )
    bp_cal = (
        qc_failures[_BP_CALIBRANT]
        if _BP_CALIBRANT in compound_cols
        else pd.Series(0, index=qc_failures.index)
    )
    plot_cal_name = aqs_to_name(_PLOT_CALIBRANT)
    bp_cal_name   = aqs_to_name(_BP_CALIBRANT)

    rows = []

    for col_codes, col_label, cal_series, cal_name in [
        (plot_cols, "All PLOT compounds", plot_cal, plot_cal_name),
        (bp_cols,   "All BP compounds",   bp_cal,   bp_cal_name),
    ]:
        # Calibrant failures — one row per merged interval for the whole column,
        # with both qualifier codes combined in the CODE cell.
        for cal_val, flag_code, bound_str in [
            (-1, "QX, LL", "below lower"),
            ( 1, "QX, LK", "above upper"),
        ]:
            cal_mask = (cal_series == cal_val).astype(int)
            reason = (
                f"{qc_name} {cal_name} (calibrant) recovery {bound_str} bound"
            )
            for start, end in compute_failure_intervals(
                all_data, cal_mask, prior_qc, next_qc
            ):
                rows.append(_make_col_row(col_label, reason, flag_code, start, end))

        # Individual compound failures when the column calibrant passed.
        qx_reason = f"{qc_name} recovery outside acceptable bounds"
        for code in col_codes:
            qx_mask = ((qc_failures[code] != 0) & (cal_series == 0)).astype(int)
            for start, end in compute_failure_intervals(
                all_data, qx_mask, prior_qc, next_qc
            ):
                rows.append(_make_row(code, qx_reason, "QX", start, end))

    if not rows:
        return pd.DataFrame(columns=_MDVR_COLUMNS)

    return _shift_and_combine(pd.DataFrame(rows))


def build_precision_qualifier_lines(
    all_data: pd.DataFrame,
    precision_failures: pd.DataFrame,
    pairs: list[tuple[pd.Timestamp, pd.Timestamp]],
    prior_qc: pd.Timestamp | None = None,
    next_qc: pd.Timestamp | None = None,
) -> pd.DataFrame:
    """Build MDVR qualifier lines for CVS precision failures.

    For each failing compound the interval spans from the hour after the
    second sample of the previous back-to-back pair to the hour before the
    first sample of the next pair.  Edge cases:

    - First pair (no previous): left bound is ``prior_qc`` or the dataset
      minimum.
    - Last pair (no next): right bound is ``next_qc`` or the dataset maximum.

    The standard ±1-hour inward shift is applied by ``_shift_and_combine``,
    so the bounds passed here are the "passing sample" timestamps, not the
    already-shifted interval edges.

    Args:
        all_data: Full Dataset.data DataFrame (for fallback interval bounds).
        precision_failures: DataFrame from check_cvs_precision.
            Index: first-run timestamp. Columns: 'filename' + integer AQS codes.
        pairs: List of (t1, t2) Timestamp tuples from check_cvs_precision.
        prior_qc: Timestamp of the last CVS precision run from the preceding
            month. Used as the left bound for the first pair. Optional.
        next_qc: Timestamp of the first CVS precision run from the following
            month. Used as the right bound for the last pair. Optional.

    Returns:
        DataFrame with MDVR qualifier columns (code QX), or empty DataFrame
        if no compounds failed.
    """
    if precision_failures.empty or not pairs:
        return pd.DataFrame(columns=_MDVR_COLUMNS)

    compound_cols = [c for c in precision_failures.columns if isinstance(c, int)]
    pair_map = {t1: i for i, (t1, _) in enumerate(pairs)}

    data_start = prior_qc if prior_qc is not None else all_data.index.min()
    data_end = next_qc if next_qc is not None else all_data.index.max()

    reason = "CVS precision outside acceptable bounds (RPD > 25%)"
    rows = []

    for ts, fail_row in precision_failures.iterrows():
        pair_idx = pair_map.get(ts)
        if pair_idx is None:
            continue

        # Left bound: 2nd sample of the previous pair.
        # _shift_and_combine will add +1h, giving "hour after prev_t2".
        left = pairs[pair_idx - 1][1] if pair_idx > 0 else data_start

        # Right bound: 1st sample of the next pair.
        # _shift_and_combine will subtract 1h, giving "hour before next_t1".
        right = pairs[pair_idx + 1][0] if pair_idx < len(pairs) - 1 else data_end

        for code in compound_cols:
            if fail_row.get(code, 0) == 1:
                rows.append(_make_row(code, reason, "QX", left, right))

    if not rows:
        return pd.DataFrame(columns=_MDVR_COLUMNS)

    return _shift_and_combine(pd.DataFrame(rows))


def build_temp_null_lines(
    temperatures: pd.Series,
    threshold: float = 30.0,
    prior_temp: pd.Timestamp | None = None,
    next_temp: pd.Timestamp | None = None,
) -> pd.DataFrame:
    """Build MDVR null lines for hours where station temperature exceeds the threshold.

    Resamples the temperature Series to hourly maximum, flags hours above
    ``threshold``, merges contiguous failure blocks into intervals, then
    applies the standard ±1-hour inward shift. All compounds are nulled for
    flagged hours (Parameter(s) = 'All Parameters').

    Args:
        temperatures: Minutely (or any frequency) temperature Series with
            DatetimeIndex, as returned by check_station_temp().temperatures.
        threshold: Station temperature threshold in °C. Default 30.0.
        prior_temp: Timestamp of the last temperature reading from the
            preceding month. Used as the left bound when the first hour of
            the month exceeds the threshold. Optional.
        next_temp: Timestamp of the first temperature reading from the
            following month. Used as the right bound when the last hour of
            the month exceeds the threshold. Optional.

    Returns:
        DataFrame with MDVR qualifier columns (code AE) ready for Excel export.
        Empty DataFrame if no hours exceed the threshold.
    """
    hourly = temperatures.resample("h").max().dropna()
    fail = (hourly > threshold).astype(int)

    if not fail.any():
        logger.info("Temperature check: no hours exceeded %.1f°C", threshold)
        return pd.DataFrame(columns=_MDVR_COLUMNS)

    proxy = pd.DataFrame(index=hourly.index)
    intervals = compute_failure_intervals(proxy, fail, prior_temp, next_temp)

    reason = f"Station temperature exceeded {threshold}\u00b0C"
    rows = []
    for start, end in intervals:
        rows.append({
            "Parameter(s)": "All Parameters",
            "COMPOUND(S) or WHOLE HOUR(S) - REASON": reason,
            "CODE": "AE",
            "startdate": pd.Timestamp(start).strftime("%m/%d/%Y"),
            "starthour": pd.Timestamp(start).strftime("%H:00"),
            "-": "-",
            "enddate": pd.Timestamp(end).strftime("%m/%d/%Y"),
            "endhour": pd.Timestamp(end).strftime("%H:00"),
            "Justification": reason,
        })

    n_hours = int(fail.sum())
    logger.info("Temperature check: %d hour(s) exceeded %.1f°C", n_hours, threshold)
    return _shift_and_combine(pd.DataFrame(rows))


# Qualifier codes that null data (written to the "Null" section of the sheet).
_NULL_CODES = frozenset({"AS", "AE"})


def _write_rows(ws, df: pd.DataFrame, start_row: int) -> None:
    """Write qualifier rows to a worksheet starting at start_row."""
    for i, (_, row) in enumerate(df.iterrows()):
        r = start_row + i
        ws[f"A{r}"] = row["Parameter(s)"]
        ws[f"D{r}"] = row["COMPOUND(S) or WHOLE HOUR(S) - REASON"]
        ws[f"I{r}"] = row["CODE"]
        ws[f"J{r}"] = row["startdate"]
        ws[f"K{r}"] = int(row["starthour"].split(":")[0])
        ws[f"L{r}"] = "-"
        ws[f"M{r}"] = row["enddate"]
        ws[f"N{r}"] = int(row["endhour"].split(":")[0])
        ws[f"O{r}"] = row["Justification"]


def write_mdvr_to_excel(
    qual_df: pd.DataFrame,
    template_path: Path,
    output_path: Path,
    null_start_row: int = 10,
    qualifier_start_row: int = 30,
) -> None:
    """Write MDVR qualifier lines to the QUALIFIERS_NULL sheet of the template.

    Rows are split by code type and written to the correct template section:
        - Null qualifiers (AS)   → "Null" section, starting at null_start_row
        - Flag qualifiers (LB, QX, ...) → "QUALIFIERS" section, starting at
          qualifier_start_row

    Column layout (matches MDVR template):
        A  Parameter(s)          (top-left of A:C merged range)
        D  Reason                (top-left of D:H merged range)
        I  CODE
        J  Start date (MM/DD/YYYY)
        K  Start hour (integer 0-23)
        L  '-'
        M  End date (MM/DD/YYYY)
        N  End hour (integer 0-23)
        O  Justification

    Args:
        qual_df: Combined DataFrame from build_blank_qualifier_lines and/or
            build_qc_qualifier_lines.
        template_path: Path to the Excel template file.
        output_path: Path for the output file. May equal template_path to
            edit in place.
        null_start_row: First data row of the Null section. Default 10.
        qualifier_start_row: First data row of the Qualifiers section. Default 29.
    """
    output_path = Path(output_path)
    if not output_path.parent.is_dir():
        logger.warning("Output directory does not exist: %s", output_path.parent)
        return

    null_df = qual_df[qual_df["CODE"].isin(_NULL_CODES)]
    flag_df = qual_df[~qual_df["CODE"].isin(_NULL_CODES)]

    wb = load_workbook(template_path)
    ws = wb["QUALIFIERS_NULL "]

    _write_rows(ws, null_df, null_start_row)
    _write_rows(ws, flag_df, qualifier_start_row)

    wb.save(output_path)
    logger.info(
        "Saved MDVR qualifiers to %s (%d null, %d flag)",
        output_path, len(null_df), len(flag_df),
    )
