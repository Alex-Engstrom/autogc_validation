# -*- coding: utf-8 -*-
"""
Fill the 'Reprocess Plan' sheet of the MDVR spreadsheet.

Reads the Dataset.data DataFrame and colour-codes each hour where a QC,
blank, or calibration sample was collected, flags missing ambient hours pink,
marks "check this hour" cells yellow, and populates the Notes cells with
overrange and daily TNMHC summaries.
"""

from calendar import monthrange
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

from autogc_validation.database.enums import SampleType

_SHEET_NAME = "Reprocess Plan"

# Column E (1-indexed 5) starts the left-panel 24-hour block.
# Column AD (30) starts the right-panel 24-hour block.
_LEFT_COL_START  = 5   # E
_RIGHT_COL_START = 30  # AD

# Row offsets relative to the "Checked PLOT" row.
_HEADER_OFFSETS      = (-2, -1)   # hour-number row, letter row (above Checked PLOT)
_COLOR_OFFSETS       = (0, 2, 4, 5)  # Checked PLOT, Checked BP, Invalid PLOT, Invalid BP
_RPO_OFFSETS         = (1, 3)    # RPO PLOT, RPO BP — always cyan "RP", applied last
_INVALID_PLOT_OFFSET = 4
_INVALID_BP_OFFSET   = 5
_NOTES_OFFSET        = 7         # Notes: merged cell row offset


# ---------------------------------------------------------------------------
# Fill colours (ARGB)
# ---------------------------------------------------------------------------
def _solid(argb: str) -> PatternFill:
    return PatternFill("solid", fgColor=argb)


_FILLS: dict[SampleType, PatternFill] = {
    SampleType.CVS:               _solid("FF0070C0"),  # blue
    SampleType.BLANK:             _solid("FF00B0F0"),  # light blue
    SampleType.LCS:               _solid("FF00B050"),  # green
    SampleType.RTS:               _solid("FFA568D2"),  # purple
    SampleType.EXPERIMENTAL:      _solid("FFFF0000"),  # red
    SampleType.CALIBRATION_POINT: _solid("FF338583"),  # teal  (Multipoint)
    SampleType.MDL_POINT:         _solid("FFFFCC99"),  # peach (Detection Limit)
}

_MISSING_FILL = _solid("FFFF99FF")  # pink  (Missing Data)
_YELLOW_FILL  = _solid("FFFFFF00")  # yellow (Check this Hour)
_RPO_FILL     = _solid("FF00FFFF")  # cyan  (Reprocessed file / RP)

# ---------------------------------------------------------------------------
# Invalid PLOT / Invalid BP text per sample type
# ---------------------------------------------------------------------------
_INVALID_TEXT: dict[SampleType, str] = {
    SampleType.CVS:               "AY",
    SampleType.BLANK:             "AY",
    SampleType.LCS:               "AY",
    SampleType.RTS:               "TC",
    SampleType.CALIBRATION_POINT: "AT",
    SampleType.EXPERIMENTAL:      "XX",
    SampleType.MDL_POINT:         "DL",
}


# ---------------------------------------------------------------------------
# Sheet structure discovery
# ---------------------------------------------------------------------------

def _discover_day_map(ws) -> dict[int, tuple[int, int]]:
    """Scan the sheet and return {day_of_month: (checked_plot_row, col_start)}.

    col_start is the 1-indexed column for hour 0 of that day's panel.
    """
    day_map: dict[int, tuple[int, int]] = {}
    for (cell,) in ws.iter_rows(min_col=3, max_col=3):
        if cell.value != "Checked PLOT":
            continue
        cp_row = cell.row
        header_row = cp_row - 3
        for col_start in (_LEFT_COL_START, _RIGHT_COL_START):
            day_num = ws.cell(row=header_row, column=col_start).value
            if isinstance(day_num, int) and 1 <= day_num <= 31:
                day_map[day_num] = (cp_row, col_start)
    return day_map


# ---------------------------------------------------------------------------
# Notes text helpers
# ---------------------------------------------------------------------------

def _strip_tz(index: pd.DatetimeIndex) -> pd.DatetimeIndex:
    return index.tz_localize(None) if getattr(index, "tz", None) is not None else index


def _build_overrange_lookup(
    overrange: pd.DataFrame | None,
) -> dict[int, list[tuple[int, str, float]]]:
    """Return {day: [(hour, compound_name, value), ...]} from overrange df."""
    if overrange is None or overrange.empty:
        return {}
    idx = _strip_tz(overrange.index)
    result: dict[int, list] = {}
    for ts, row in zip(idx, overrange.itertuples()):
        result.setdefault(ts.day, []).append((ts.hour, row.compound_name, float(row.value)))
    return result


def _build_tnmhc_lookup(
    daily_tnmhc: pd.Series | None,
) -> dict[int, tuple[int, float]]:
    """Return {day: (hour, tnmhc_value)} from daily_tnmhc series."""
    if daily_tnmhc is None or daily_tnmhc.empty:
        return {}
    idx = _strip_tz(daily_tnmhc.index)
    return {ts.day: (ts.hour, float(val)) for ts, val in zip(idx, daily_tnmhc.values)}


def _format_notes(
    day: int,
    overrange_by_day: dict[int, list],
    tnmhc_by_day: dict[int, tuple],
) -> str:
    """Build the Notes cell text for a single day."""
    lines = []

    # Group overrange entries by hour for compactness.
    by_hour: dict[int, list[str]] = {}
    for hour, cname, val in sorted(overrange_by_day.get(day, []), key=lambda x: (x[0], x[1])):
        by_hour.setdefault(hour, []).append(f"{cname} {val:.2f}")
    for hour, entries in sorted(by_hour.items()):
        lines.append(f"Overrange ({hour:02d}:00): {', '.join(entries)} ppbC")

    if day in tnmhc_by_day:
        hour, val = tnmhc_by_day[day]
        lines.append(f"Daily max TNMHC: {val:.1f} ppbC ({hour:02d}:00)")

    return "\n".join(lines)


# ---------------------------------------------------------------------------
# Main function
# ---------------------------------------------------------------------------

def fill_reprocess_plan(
    data_df: pd.DataFrame,
    template_path: Path,
    output_path: Path,
    year: int,
    month: int,
    overrange: pd.DataFrame | None = None,
    daily_tnmhc: pd.Series | None = None,
) -> None:
    """Colour-code sample hours in the Reprocess Plan sheet.

    Fills are applied in the following order so that higher-priority layers
    paint over lower-priority ones:

    1. **Missing data (pink)** on the four non-RPO data rows for every hour
       with no entry in *data_df*.
    2. **Yellow "Check this Hour"** on the two header rows (hour-number and
       letter) for:

       - Each QC/blank/cal sample hour.
       - The hour immediately following each QC/blank/cal sample run.
       - Hour 0 of the first day and hour 23 of the last day of the month.
       - Every hour flagged by *overrange* or *daily_tnmhc*.

    3. **Sample-type colour** on the four non-RPO data rows for QC/blank/cal
       hours, plus the appropriate qualifier text in Invalid PLOT / Invalid BP.
    4. **Cyan "RP"** on *all* 24 columns of the RPO PLOT and RPO BP rows for
       every day block — always applied last.

    The Notes merged cell beneath each day's grid is populated with a summary
    of any overrange exceedances and the daily TNMHC maximum for that day.

    Args:
        data_df: Dataset.data concentration DataFrame — DatetimeIndex with
            sample timestamps and a ``sample_type`` column.
        template_path: Path to the MDVR .xlsx file to read.
        output_path: Path to write the modified workbook. May equal
            *template_path* to modify in-place.
        year: Validation month year.
        month: Validation month number (1-12).
        overrange: Optional DataFrame returned by ``check_overrange_values``.
            Columns: compound (int), value (float), compound_name (str).
        daily_tnmhc: Optional Series returned by ``check_daily_max_tnmhc``,
            indexed by the timestamp of each day's maximum TNMHC sample.
    """
    template_path = Path(template_path)
    output_path   = Path(output_path)

    wb = load_workbook(template_path)
    if _SHEET_NAME not in wb.sheetnames:
        raise ValueError(f"Sheet '{_SHEET_NAME}' not found in {template_path}")
    ws = wb[_SHEET_NAME]

    day_map = _discover_day_map(ws)
    if not day_map:
        raise RuntimeError(
            f"Could not locate day blocks in '{_SHEET_NAME}'. "
            "Verify the sheet structure matches the expected template."
        )

    n_days = monthrange(year, month)[1]

    # Exclude any extra day-31 panels present in templates for shorter months.
    day_map = {d: v for d, v in day_map.items() if d <= n_days}

    # Strip timezone from data_df index.
    data_index = _strip_tz(data_df.index)

    # ------------------------------------------------------------------
    # Build lookup structures
    # ------------------------------------------------------------------

    existing_hours: set[tuple[int, int]] = {(ts.day, ts.hour) for ts in data_index}

    target_types = set(_FILLS)
    qc_hours: dict[tuple[int, int], SampleType] = {}
    for ts, st_val in zip(data_index, data_df["sample_type"]):
        if st_val in target_types:
            qc_hours[(ts.day, ts.hour)] = SampleType(st_val)

    overrange_by_day = _build_overrange_lookup(overrange)
    tnmhc_by_day     = _build_tnmhc_lookup(daily_tnmhc)

    # Hours that receive yellow header treatment.
    yellow_full: set[tuple[int, int]] = set()

    # QC/blank/cal sample hours themselves.
    for day, hour in qc_hours:
        yellow_full.add((day, hour))

    # Hour immediately following each QC/blank/cal run.
    for day, hour in qc_hours:
        if hour < 23:
            yellow_full.add((day, hour + 1))

    # First hour of the first day and last hour of the last day.
    yellow_full.add((1, 0))
    yellow_full.add((n_days, 23))

    # Overrange and daily TNMHC max hours.
    overrange_index = _strip_tz(overrange.index) if overrange is not None and not overrange.empty else []
    for ts in overrange_index:
        yellow_full.add((ts.day, ts.hour))

    tnmhc_index = _strip_tz(daily_tnmhc.index) if daily_tnmhc is not None and not daily_tnmhc.empty else []
    for ts in tnmhc_index:
        yellow_full.add((ts.day, ts.hour))

    # ------------------------------------------------------------------
    # Apply fills
    # ------------------------------------------------------------------

    missing_count = sample_count = 0

    for day, (cp_row, col_start) in day_map.items():

        # Step 1 — missing data (pink) on non-RPO data rows.
        for hour in range(24):
            if (day, hour) not in existing_hours:
                col = col_start + hour
                for offset in _COLOR_OFFSETS:
                    ws.cell(row=cp_row + offset, column=col).fill = _MISSING_FILL
                missing_count += 1

        # Step 2 — yellow header rows for check hours.
        for _, hour in [(d, h) for d, h in yellow_full if d == day]:
            col = col_start + hour
            for offset in _HEADER_OFFSETS:
                ws.cell(row=cp_row + offset, column=col).fill = _YELLOW_FILL

        # Step 3 — sample-type colour on data rows + invalid text.
        for (d, hour), sample_type in qc_hours.items():
            if d != day:
                continue
            col          = col_start + hour
            fill         = _FILLS[sample_type]
            invalid_text = _INVALID_TEXT[sample_type]
            for offset in _COLOR_OFFSETS:
                ws.cell(row=cp_row + offset, column=col).fill = fill
            ws.cell(row=cp_row + _INVALID_PLOT_OFFSET, column=col).value = invalid_text
            ws.cell(row=cp_row + _INVALID_BP_OFFSET,   column=col).value = invalid_text
            sample_count += 1

        # Step 4 — RPO rows: cyan "RP" across all 24 hours (applied last).
        for hour in range(24):
            col = col_start + hour
            for offset in _RPO_OFFSETS:
                cell = ws.cell(row=cp_row + offset, column=col)
                cell.fill  = _RPO_FILL
                cell.value = "RP"

        # Notes cell — overrange and TNMHC summary for this day.
        notes_text = _format_notes(day, overrange_by_day, tnmhc_by_day)
        if notes_text:
            ws.cell(row=cp_row + _NOTES_OFFSET, column=col_start).value = notes_text

    wb.save(output_path)
    print(
        f"Reprocess Plan filled: {sample_count} QC/blank hour(s) coloured, "
        f"{missing_count} missing hour(s) flagged pink."
    )
