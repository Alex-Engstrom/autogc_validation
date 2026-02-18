# -*- coding: utf-8 -*-
"""
MDVR (Monthly Data Validation Report) qualifier generation.

Computes failure intervals from pass/fail time series, builds
qualifier lines for blank and QC failures, and writes them to
Excel templates for upload.
"""

import logging
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import load_workbook

from autogc_validation.database.enums import aqs_to_name

logger = logging.getLogger(__name__)

QC_TYPE_NAMES = {
    "c": "CVS",
    "e": "LCS",
    "q": "RTS",
}

_MDVR_COLUMNS = [
    "Parameter(s)",
    "COMPOUND(S) or WHOLE HOUR(S) - REASON",
    "CODE",
    "startdate",
    "starthour",
    "-",
    "enddate",
    "endhour",
    "Justification",
]


def compute_failure_intervals(
    all_data: pd.DataFrame,
    failure_series: pd.Series,
) -> np.ndarray:
    """Compute merged failure intervals from a binary pass/fail series.

    Failures are bounded by the nearest passing observations on each side.
    Overlapping intervals are merged.

    Args:
        all_data: Full dataset DataFrame (used for min/max time bounds).
        failure_series: Series with datetime index and 0/1 values (1 = fail).

    Returns:
        Nx2 numpy array of [start_time, end_time] for each merged interval.
    """
    failure_series = failure_series.copy().sort_index()
    idx = failure_series.index
    lower_bound = all_data.index.min()
    upper_bound = all_data.index.max()

    passes = failure_series == 0
    fails = failure_series == 1

    if not fails.any():
        return np.empty((0, 2), dtype="datetime64[ns]")

    pass_idx = pd.Series(idx.where(passes, pd.NaT), index=idx)

    left_bounds = pass_idx.ffill().fillna(lower_bound)
    right_bounds = pass_idx.bfill().fillna(upper_bound)

    lb = left_bounds[fails].to_numpy()
    rb = right_bounds[fails].to_numpy()

    intervals = np.column_stack([lb, rb])
    intervals = intervals[np.argsort(intervals[:, 0])]

    lb = intervals[:, 0]
    rb = intervals[:, 1]

    # Merge overlapping intervals
    new_interval_mask = np.r_[True, lb[1:] > np.maximum.accumulate(rb[:-1])]
    interval_ids = np.cumsum(new_interval_mask) - 1

    merged = np.zeros((interval_ids.max() + 1, 2), dtype=intervals.dtype)
    merged[:, 0] = np.minimum.reduceat(lb, np.flatnonzero(new_interval_mask))
    merged[:, 1] = np.maximum.reduceat(rb, np.flatnonzero(new_interval_mask))

    return merged


def _shift_and_combine(df: pd.DataFrame) -> pd.DataFrame:
    """Group qualifier rows by time interval, combine compounds, shift times."""
    df_combined = (
        df.groupby(["startdate", "starthour", "enddate", "endhour"])
        .agg({
            "Parameter(s)": lambda x: ", ".join(x),
            "COMPOUND(S) or WHOLE HOUR(S) - REASON": "first",
            "CODE": "first",
            "-": "first",
            "Justification": "first",
        })
        .reset_index()
    )

    # Parse into datetimes, shift by ±1 hour, split back
    df_combined["start_dt"] = pd.to_datetime(
        df_combined["startdate"] + " " + df_combined["starthour"],
        format="%m/%d/%Y %H:%M",
    )
    df_combined["end_dt"] = pd.to_datetime(
        df_combined["enddate"] + " " + df_combined["endhour"],
        format="%m/%d/%Y %H:%M",
    )

    df_combined["start_dt"] += pd.Timedelta(hours=1)
    df_combined["end_dt"] -= pd.Timedelta(hours=1)

    df_combined["startdate"] = df_combined["start_dt"].dt.strftime("%m/%d/%Y")
    df_combined["starthour"] = df_combined["start_dt"].dt.strftime("%H:00")
    df_combined["enddate"] = df_combined["end_dt"].dt.strftime("%m/%d/%Y")
    df_combined["endhour"] = df_combined["end_dt"].dt.strftime("%H:00")

    return df_combined.drop(columns=["start_dt", "end_dt"])


def build_blank_qualifier_lines(
    all_data: pd.DataFrame,
    blank_failures_wide: pd.DataFrame,
) -> pd.DataFrame:
    """Build MDVR qualifier lines for blank failures.

    Args:
        all_data: Full Dataset.data DataFrame (for interval bounds).
        blank_failures_wide: Wide-format output from
            qc.blanks.compounds_above_mdl_wide().

    Returns:
        DataFrame with MDVR columns ready for Excel export.
    """
    rows = []
    for compound_code in blank_failures_wide.columns:
        merged = compute_failure_intervals(
            all_data, blank_failures_wide[compound_code]
        )
        for start, end in merged:
            rows.append({
                "Parameter(s)": aqs_to_name(compound_code),
                "COMPOUND(S) or WHOLE HOUR(S) - REASON": "Blank(s) above respective MDL(s)",
                "CODE": "LB",
                "startdate": start.strftime("%m/%d/%Y"),
                "starthour": start.strftime("%H:00"),
                "-": "-",
                "enddate": end.strftime("%m/%d/%Y"),
                "endhour": end.strftime("%H:00"),
                "Justification": "Blank(s) above respective MDL(s)",
            })

    if not rows:
        return pd.DataFrame(columns=_MDVR_COLUMNS)

    return _shift_and_combine(pd.DataFrame(rows))


def build_qc_qualifier_lines(
    all_data: pd.DataFrame,
    qc_failures_wide: pd.DataFrame,
    qc_type: str,
) -> pd.DataFrame:
    """Build MDVR qualifier lines for QC recovery failures.

    Args:
        all_data: Full Dataset.data DataFrame.
        qc_failures_wide: Wide-format output from
            qc.recovery.check_qc_recovery_wide().
        qc_type: Sample type code — 'c', 'e', or 'q'.

    Returns:
        DataFrame with MDVR columns.
    """
    qc_name = QC_TYPE_NAMES[qc_type]
    reason = f"{qc_name} recovery outside acceptable bounds"

    rows = []
    for compound_code in qc_failures_wide.columns:
        merged = compute_failure_intervals(
            all_data, qc_failures_wide[compound_code]
        )
        for start, end in merged:
            rows.append({
                "Parameter(s)": aqs_to_name(compound_code),
                "COMPOUND(S) or WHOLE HOUR(S) - REASON": reason,
                "CODE": "QX",
                "startdate": start.strftime("%m/%d/%Y"),
                "starthour": start.strftime("%H:00"),
                "-": "-",
                "enddate": end.strftime("%m/%d/%Y"),
                "endhour": end.strftime("%H:00"),
                "Justification": reason,
            })

    if not rows:
        return pd.DataFrame(columns=_MDVR_COLUMNS)

    return _shift_and_combine(pd.DataFrame(rows))


def write_mdvr_to_excel(
    qual_df: pd.DataFrame,
    template_path: Path,
    output_path: Path,
) -> None:
    """Write MDVR qualifier lines to an Excel template.

    Args:
        qual_df: DataFrame from build_blank_qualifier_lines or
            build_qc_qualifier_lines.
        template_path: Path to the Excel template file.
        output_path: Path for the output Excel file.
    """
    output_path = Path(output_path)
    if not output_path.parent.is_dir():
        logger.warning("Output directory does not exist: %s", output_path.parent)
        return

    wb = load_workbook(template_path)
    ws = wb.active

    for i, (_, row) in enumerate(qual_df.iterrows()):
        ws[f"A{i + 1}"] = row["Parameter(s)"]
        ws[f"D{i + 1}"] = row["COMPOUND(S) or WHOLE HOUR(S) - REASON"]
        ws[f"I{i + 1}"] = row["CODE"]
        ws[f"J{i + 1}"] = row["startdate"]
        ws[f"K{i + 1}"] = row["starthour"]
        ws[f"M{i + 1}"] = row["enddate"]
        ws[f"N{i + 1}"] = row["endhour"]
        ws[f"O{i + 1}"] = row["Justification"]

    wb.save(output_path)
    logger.info("Saved MDVR output to %s", output_path)
